import openpyxl
from tkinter import messagebox
from db_operations import insertar_datos

def procesar_archivo(ruta_archivo):
    try:
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)

        for hoja in wb.sheetnames:
            hoja_actual = wb[hoja]
            datos = []
            columnas_por_tabla = 7  
            
            for inicio_columna in range(1, hoja_actual.max_column, columnas_por_tabla):
                mes = None
            
                mes_fila = hoja_actual.cell(row=2, column=inicio_columna + 1).value
                if not mes_fila:
                    mes_fila = hoja_actual.cell(row=3, column=inicio_columna + 1).value
                if mes_fila:
                    mes = mes_fila   
                
                for idx, fila in enumerate(hoja_actual.iter_rows(min_row=4, min_col=inicio_columna +1, max_col=inicio_columna +5, values_only=True)):
                                                                
                    if all(fila):
                        datos.append((
                            fila[0],  
                            fila[1],  
                            fila[2],  
                            fila[3],  
                            fila[4],  
                            mes       
                        ))

                if datos:
                    insertar_datos(datos)
                    datos = []  
            
        messagebox.showinfo("Éxito", "Datos insertados correctamente")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo procesar el archivo: {e}")